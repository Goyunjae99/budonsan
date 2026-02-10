"""
엑셀 파일 저장 기능
"""

import pandas as pd
from datetime import datetime
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def save_to_excel(data: List[Dict[str, str]], filename: str) -> bool:
    """
    데이터를 엑셀 파일로 저장
    
    Args:
        data: 저장할 데이터 리스트
        filename: 저장할 파일 경로
        
    Returns:
        성공 여부
    """
    try:
        if not data:
            return False
        
        # DataFrame 생성
        df = pd.DataFrame(data)
        
        # 컬럼 순서 지정
        columns = ['동', '가격', '면적', '층수']
        df = df.reindex(columns=columns)
        
        # 엑셀 파일로 저장
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='매물정보', index=False)
            
            # 워크시트 가져오기
            worksheet = writer.sheets['매물정보']
            
            # 헤더 스타일 설정
            header_font = Font(bold=True, size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            for cell in worksheet[1]:
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 데이터 정렬 설정
            data_alignment = Alignment(horizontal='left', vertical='center')
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                for cell in row:
                    cell.alignment = data_alignment
            
            # 컬럼 너비 자동 조정
            column_widths = {
                'A': 15,  # 동
                'B': 20,  # 가격
                'C': 15,  # 면적
                'D': 15   # 층수
            }
            
            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width
        
        return True
        
    except Exception as e:
        print(f"엑셀 파일 저장 중 오류: {e}")
        return False


def generate_default_filename(prefix: str = "센트럴파크_매물정보") -> str:
    """
    기본 파일명 생성
    
    Args:
        prefix: 파일명 접두사
        
    Returns:
        생성된 파일명
    """
    now = datetime.now()
    timestamp = now.strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{timestamp}.xlsx"
